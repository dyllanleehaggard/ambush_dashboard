import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
import numpy as np
import openpyxl
from datetime import datetime
import os

# ──────────────────────────────────────────────
# PAGE CONFIG
# ──────────────────────────────────────────────
st.set_page_config(
    page_title="STL Ambush — Coaching Intelligence",
    page_icon="STL",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ──────────────────────────────────────────────
# CUSTOM CSS
# ──────────────────────────────────────────────
TEAL = "#077988"
TEAL_LIGHT = "#0a9fb3"
TEAL_DARK = "#055f6b"
BG = "#0a0a0a"
CARD_BG = "#111111"
CARD_BORDER = "#1a1a1a"

st.markdown(f"""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Anton&display=swap');

    /* Force dark everywhere */
    .stApp, .main, [data-testid="stAppViewContainer"],
    [data-testid="stHeader"], section[data-testid="stSidebar"] {{
        background-color: {BG} !important;
        color: #e0e0e0 !important;
    }}
    /* Kill Streamlit's colored header decoration line */
    [data-testid="stHeader"] {{
        background: {BG} !important;
    }}
    [data-testid="stDecoration"] {{
        display: none !important;
    }}
    /* Remove top padding from main content */
    .block-container {{
        padding-top: 0 !important;
    }}
    [data-testid="stAppViewContainer"] > .main {{
        padding-top: 0 !important;
    }}
    /* Hide sidebar completely */
    [data-testid="stSidebar"] {{
        display: none !important;
    }}
    [data-testid="collapsedControl"] {{
        display: none !important;
    }}
    button[kind="header"] {{
        display: none !important;
    }}

    /* Impact-style headlines via Anton */
    h1, h2, h3 {{
        font-family: 'Anton', 'Impact', sans-serif !important;
        text-transform: uppercase !important;
        letter-spacing: 1.5px !important;
    }}
    h1 {{ color: #ffffff !important; font-size: 1.8rem !important; }}
    h2 {{ color: {TEAL} !important; font-size: 1.2rem !important; }}
    h3 {{ color: #cccccc !important; font-size: 0.95rem !important; }}

    /* Body text in Arial */
    p, span, div, li, td, th, label, .stMarkdown {{
        font-family: Arial, Helvetica, sans-serif !important;
    }}

    /* Metric cards */
    [data-testid="stMetric"] {{
        background: {CARD_BG};
        border: 1px solid {CARD_BORDER};
        border-left: 3px solid {TEAL};
        padding: 8px 10px;
        border-radius: 6px;
        overflow: hidden;
    }}
    [data-testid="stMetric"] > div {{
        overflow: hidden;
    }}
    [data-testid="stMetricLabel"] {{
        color: #999 !important;
        font-size: 0.6rem !important;
        text-transform: uppercase;
        letter-spacing: 0.3px;
        white-space: nowrap;
        overflow: hidden;
        text-overflow: ellipsis;
    }}
    [data-testid="stMetricValue"] {{
        color: #ffffff !important;
        font-family: 'Anton', 'Impact', sans-serif !important;
        font-size: 1.05rem !important;
        white-space: nowrap;
        overflow: hidden;
        text-overflow: ellipsis;
        line-height: 1.3 !important;
    }}
    [data-testid="stMetricDelta"] > div {{
        font-size: 0.6rem !important;
        white-space: nowrap;
        overflow: hidden;
        text-overflow: ellipsis;
    }}

    /* Tabs */
    .stTabs [data-baseweb="tab-list"] {{
        gap: 0px;
        background: #111;
        border-radius: 8px;
        padding: 4px;
    }}
    .stTabs [data-baseweb="tab"] {{
        background: transparent !important;
        color: #888 !important;
        border-radius: 6px;
        padding: 8px 20px;
        font-family: 'Anton', 'Impact', sans-serif !important;
        text-transform: uppercase;
        letter-spacing: 1px;
        font-size: 0.7rem;
    }}
    .stTabs [aria-selected="true"] {{
        background: {TEAL} !important;
        color: #fff !important;
    }}

    /* Tables */
    .stDataFrame {{ border-radius: 8px; overflow: hidden; }}
    .stDataFrame table {{ background: {CARD_BG} !important; }}
    .stDataFrame th {{
        background: {TEAL_DARK} !important;
        color: white !important;
        font-family: 'Anton', 'Impact', sans-serif !important;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }}
    .stDataFrame td {{ color: #d0d0d0 !important; border-color: #222 !important; }}

    /* Selectbox / inputs */
    .stSelectbox > div > div,
    .stMultiSelect > div > div {{
        background: #1a1a1a !important;
        border-color: #333 !important;
        color: #e0e0e0 !important;
    }}

    /* Expander */
    .streamlit-expanderHeader {{
        background: {CARD_BG} !important;
        border: 1px solid #222 !important;
        border-radius: 6px !important;
        color: {TEAL_LIGHT} !important;
        font-family: 'Anton', 'Impact', sans-serif !important;
        text-transform: uppercase;
    }}

    /* Divider */
    hr {{ border-color: #222 !important; }}

    /* Custom classes */
    .stat-card {{
        background: {CARD_BG};
        border: 1px solid {CARD_BORDER};
        border-radius: 8px;
        padding: 20px;
        margin: 8px 0;
    }}
    .streak-legend {{
        display: flex;
        gap: 12px;
        justify-content: center;
        font-size: 0.65rem;
        color: #888;
        margin-top: -10px;
    }}

    /* Hide streamlit branding */
    #MainMenu {{ visibility: hidden; }}
    footer {{ visibility: hidden; }}
    header {{ visibility: hidden; }}

    /* Force all text */
    * {{ color-scheme: dark; }}

    /* Global font size reduction */
    .stApp {{ font-size: 0.82rem !important; }}
    [data-testid="stMarkdownContainer"] p {{
        font-size: 0.82rem !important;
    }}
    .stSelectbox label, .stMultiSelect label, .stTextInput label {{
        font-size: 0.75rem !important;
        text-transform: uppercase;
        letter-spacing: 0.3px;
        color: #888 !important;
    }}
    .stRadio label {{
        font-size: 0.78rem !important;
    }}
    .stDataFrame td, .stDataFrame th {{
        font-size: 0.72rem !important;
        padding: 4px 8px !important;
    }}
    .stPlotlyChart {{
        margin-top: -8px;
        margin-bottom: -8px;
    }}
    [data-testid="stHorizontalBlock"] {{
        gap: 0.4rem !important;
    }}
</style>
""", unsafe_allow_html=True)

# ──────────────────────────────────────────────
# DATA LOADING
# ──────────────────────────────────────────────
@st.cache_data
def load_data():
    possible_paths = [
        "/mnt/user-data/uploads/Historical.xlsx",
        "Historical.xlsx",
        os.path.join(os.path.dirname(__file__), "Historical.xlsx")
    ]
    xlsx_path = None
    for p in possible_paths:
        if os.path.exists(p):
            xlsx_path = p
            break
    if xlsx_path is None:
        st.error("Cannot find Historical.xlsx")
        st.stop()

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    data = {}

    # ── PLAYER STATS ──
    ws = wb['Player']
    rows = list(ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=15, values_only=True))
    player_records = []
    for r in rows:
        if r[1] is not None and r[0] is not None:
            player_records.append({
                '#': int(r[0]) if r[0] else 0,
                'Name': r[1], 'Pos': r[2],
                'G': r[3] or 0, 'A': r[4] or 0, 'Pts': r[5] or 0,
                'PIM': r[6] or 0, 'F': r[7] or 0, 'SOG': r[8] or 0,
                'BS': r[9] or 0, 'BC': r[10] or 0, 'YC': r[11] or 0,
                'RC': r[12] or 0, 'Sh': r[13] or 0, 'GP': r[14] or 0
            })
    data['players'] = pd.DataFrame(player_records)
    if len(data['players']) > 0:
        df = data['players']
        df['G/GP'] = (df['G'] / df['GP'].replace(0, np.nan)).round(2)
        df['Pts/GP'] = (df['Pts'] / df['GP'].replace(0, np.nan)).round(2)
        df['SOG%'] = ((df['SOG'] / df['Sh'].replace(0, np.nan)) * 100).round(1)
        df['Sh/GP'] = (df['Sh'] / df['GP'].replace(0, np.nan)).round(1)

    # ── GOALKEEPER STATS ──
    ws = wb['Goalkeeper']
    rows = list(ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=13, values_only=True))
    gk_records = []
    for r in rows:
        if r[1] is not None and r[0] is not None:
            gk_records.append({
                '#': int(r[0]) if r[0] else 0,
                'Name': r[1],
                'SA': r[2] or 0, 'GA': r[3] or 0, 'Sv': r[4] or 0,
                'Sv%': round(r[5] * 100, 1) if r[5] else 0,
                'PIM': r[6] or 0, 'G': r[7] or 0, 'A': r[8] or 0,
                'F': r[9] or 0, 'BC': r[10] or 0, 'YC': r[11] or 0,
                'RC': r[12] or 0
            })
    data['goalkeepers'] = pd.DataFrame(gk_records)

    # ── PLUS MINUS ──
    ws = wb['PlusMinus']
    header_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    game_dates_raw = list(header_row[6:25])
    game_dates = [str(d) for d in game_dates_raw if d is not None]

    rows = list(ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=25, values_only=True))
    pm_records = []
    for r in rows:
        if r[0] is not None and isinstance(r[0], str):
            game_vals = []
            for v in r[6:6+len(game_dates)]:
                if v is None or v == '-' or v == '':
                    game_vals.append(None)
                else:
                    try:
                        game_vals.append(float(v))
                    except:
                        game_vals.append(None)
            pm_records.append({
                'Name': r[0],
                'Plus': r[1] or 0, 'Minus': r[2] or 0,
                'Total': r[3] or 0,
                'Trend': str(r[4]) if r[4] is not None else '-',
                'game_values': game_vals
            })
    data['plusminus'] = pm_records
    data['pm_game_dates'] = game_dates

    # ── HUDL ASSIST DATA ──
    ws = wb['Hudl Assist Data']
    headers = [str(c.value).strip() if c.value else '' for c in list(ws.iter_rows(min_row=1, max_row=1))[0][:30]]
    rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=30, values_only=True))
    hudl_records = []
    for r in rows:
        if r[0] is not None and r[1] is not None and isinstance(r[0], datetime):
            rec = {'Date': r[0], 'Team': r[1]}
            for i, h in enumerate(headers[2:], start=2):
                if h and i < len(r):
                    rec[h] = r[i] if r[i] is not None else 0
            hudl_records.append(rec)
    data['hudl'] = pd.DataFrame(hudl_records)
    if len(data['hudl']) > 0:
        data['hudl']['Date'] = pd.to_datetime(data['hudl']['Date'])
        data['hudl']['Date_str'] = data['hudl']['Date'].dt.strftime('%m/%d/%Y')

    return data


data = load_data()

# ──────────────────────────────────────────────
# PLOTLY THEME
# ──────────────────────────────────────────────
PLOTLY_LAYOUT = dict(
    paper_bgcolor='rgba(0,0,0,0)',
    plot_bgcolor='rgba(17,17,17,0.8)',
    font=dict(family='Arial, sans-serif', color='#ccc', size=10),
    xaxis=dict(gridcolor='#222', zerolinecolor='#333'),
    yaxis=dict(gridcolor='#222', zerolinecolor='#333'),
    margin=dict(l=40, r=20, t=40, b=40),
    hoverlabel=dict(bgcolor='#1a1a1a', font_color='#fff', bordercolor=TEAL),
)


def styled_fig(fig, height=400):
    fig.update_layout(**PLOTLY_LAYOUT, height=height)
    return fig


# ──────────────────────────────────────────────
# TOP NAVIGATION BAR
# ──────────────────────────────────────────────
NAV_PAGES = ["Dashboard Home", "Plus/Minus", "Hudl Trend Analysis", "Player Stats", "Contact"]

if 'page' not in st.session_state:
    st.session_state['page'] = "Dashboard Home"

# Brand header
st.markdown(f"""
<div style="background:#080808; padding:14px 0 8px 0; margin:0 -1rem 0 -1rem; text-align:center;">
    <span style="font-family:'Anton',Impact,sans-serif; font-size:1.3rem; color:{TEAL}; letter-spacing:3px;">STL</span>
    <span style="font-family:'Anton',Impact,sans-serif; font-size:1.3rem; color:white; letter-spacing:3px;"> AMBUSH</span>
    <span style="font-family:Arial,sans-serif; font-size:0.5rem; color:#555; letter-spacing:2px; display:block; margin-top:2px;">COACHING INTELLIGENCE PLATFORM</span>
</div>
""", unsafe_allow_html=True)

# Navigation buttons
nav_cols = st.columns(len(NAV_PAGES))
for i, p in enumerate(NAV_PAGES):
    with nav_cols[i]:
        is_active = st.session_state['page'] == p
        if st.button(p, key=f"nav_{p}", use_container_width=True, type="primary" if is_active else "secondary"):
            st.session_state['page'] = p
            st.rerun()

st.markdown(f"""
<style>
    [data-testid="stHorizontalBlock"]:first-of-type {{
        background: #080808;
        margin: 0 -1rem 0.5rem -1rem;
        padding: 0 12px 10px 12px;
        border-bottom: 2px solid #1a1a1a;
        gap: 4px !important;
    }}
    [data-testid="stHorizontalBlock"]:first-of-type button {{
        font-family: 'Anton', 'Impact', sans-serif !important;
        text-transform: uppercase !important;
        letter-spacing: 1px !important;
        font-size: 0.65rem !important;
        border-radius: 6px !important;
        padding: 6px 8px !important;
        min-height: 0 !important;
        height: auto !important;
        line-height: 1.3 !important;
        white-space: nowrap !important;
    }}
    [data-testid="stHorizontalBlock"]:first-of-type button[kind="primary"] {{
        background: {TEAL} !important;
        color: #fff !important;
        border: none !important;
    }}
    [data-testid="stHorizontalBlock"]:first-of-type button[kind="secondary"] {{
        background: transparent !important;
        color: #777 !important;
        border: 1px solid #333 !important;
    }}
    [data-testid="stHorizontalBlock"]:first-of-type button[kind="secondary"]:hover {{
        color: #ccc !important;
        background: #151515 !important;
    }}
    @media (max-width: 768px) {{
        [data-testid="stHorizontalBlock"]:first-of-type button {{
            font-size: 0.5rem !important;
            padding: 5px 4px !important;
            letter-spacing: 0.3px !important;
        }}
    }}
</style>
""", unsafe_allow_html=True)

page = st.session_state['page']


# ══════════════════════════════════════════════
# PAGE: DASHBOARD HOME
# ══════════════════════════════════════════════
if page == "Dashboard Home":
    st.markdown(f"""
    <h1 style="text-align:center; letter-spacing:4px; margin-bottom:0;">
        ST. LOUIS AMBUSH
    </h1>
    <p style="text-align:center; color:{TEAL}; font-family:'Anton',Impact,sans-serif; font-size:0.9rem; letter-spacing:3px; margin-top:0;">
        COACHING INTELLIGENCE DASHBOARD
    </p>
    """, unsafe_allow_html=True)

    st.markdown("---")

    players = data['players']
    hudl = data['hudl']
    ambush_hudl = hudl[hudl['Team'].str.contains('Ambush', case=False, na=False)]

    total_goals = int(players['G'].sum())
    total_assists = int(players['A'].sum())
    games_in_hudl = len(ambush_hudl)

    gk = data['goalkeepers']
    total_sa = gk['SA'].sum()
    total_sv = gk['Sv'].sum()
    sv_pct = round((total_sv / total_sa) * 100, 1) if total_sa > 0 else 0

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("RECORD", "12-5-5")
    c2.metric("GOALS SCORED", total_goals)
    c3.metric("POSITION", "#5")
    c4.metric("GK SAVE %", f"{sv_pct}%")
    c5.metric("GAMES TRACKED", games_in_hudl)

    st.markdown("---")

    # Two-column layout
    col_left, col_right = st.columns([1.2, 1])

    with col_left:
        st.markdown("## TOP SCORERS")
        top_scorers = players.nlargest(8, 'Pts')[['#', 'Name', 'Pos', 'G', 'A', 'Pts', 'GP', 'Pts/GP']].reset_index(drop=True)

        fig = go.Figure()
        fig.add_trace(go.Bar(
            y=top_scorers['Name'], x=top_scorers['G'],
            name='Goals', orientation='h',
            marker_color=TEAL, text=top_scorers['G'],
            textposition='inside', textfont=dict(color='white', size=9)
        ))
        fig.add_trace(go.Bar(
            y=top_scorers['Name'], x=top_scorers['A'],
            name='Assists', orientation='h',
            marker_color='#444', text=top_scorers['A'],
            textposition='inside', textfont=dict(color='white', size=9)
        ))
        fig.update_layout(barmode='stack', yaxis=dict(autorange='reversed'))
        fig = styled_fig(fig, 350)
        fig.update_layout(
            legend=dict(orientation='h', y=1.12, x=0.5, xanchor='center', font=dict(color='#aaa')),
            yaxis_title=None, xaxis_title='Points Contribution'
        )
        st.plotly_chart(fig, use_container_width=True)

    with col_right:
        st.markdown("## PLUS/MINUS LEADERS")
        pm_data = data['plusminus']
        pm_df = pd.DataFrame([{
            'Name': p['Name'], 'Total': p['Total']
        } for p in pm_data]).sort_values('Total', ascending=False).head(10)

        fig = go.Figure()
        colors = [TEAL if v >= 0 else '#cc3333' for v in pm_df['Total']]
        fig.add_trace(go.Bar(
            x=pm_df['Name'], y=pm_df['Total'],
            marker_color=colors,
            text=pm_df['Total'].apply(lambda x: f"+{int(x)}" if x > 0 else str(int(x))),
            textposition='outside', textfont=dict(color='white', size=9)
        ))
        fig.add_hline(y=0, line_dash="dash", line_color="#555")
        fig = styled_fig(fig, 350)
        fig.update_layout(xaxis_tickangle=-35, yaxis_title='+/-')
        st.plotly_chart(fig, use_container_width=True)

    st.markdown("---")

    # Season Progression
    st.markdown("## SEASON PROGRESSION — KEY METRICS")
    if len(ambush_hudl) > 0:
        amb_sorted = ambush_hudl.sort_values('Date')
        col_map = {c.strip(): c for c in amb_sorted.columns}

        def plot_sparkline(container, col_name, label):
            real_col = None
            for c in amb_sorted.columns:
                if c.strip() == col_name.strip():
                    real_col = c
                    break
            if real_col and real_col in amb_sorted.columns:
                with container:
                    vals = amb_sorted[real_col].values
                    avg_val = np.mean(vals)
                    fig = go.Figure()
                    fig.add_trace(go.Scatter(
                        x=amb_sorted['Date_str'], y=vals,
                        mode='lines+markers',
                        line=dict(color=TEAL, width=2),
                        marker=dict(color=TEAL, size=6),
                        fill='tozeroy',
                        fillcolor='rgba(7,121,136,0.15)'
                    ))
                    fig.add_hline(y=avg_val, line_dash="dash", line_color="#666", annotation_text="avg")
                    fig = styled_fig(fig, 200)
                    fig.update_layout(
                        showlegend=False,
                        xaxis=dict(showticklabels=False),
                        margin=dict(l=30, r=10, t=30, b=10),
                        title=dict(text=label, font=dict(size=10, color='#999'))
                    )
                    st.plotly_chart(fig, use_container_width=True)

        row1 = st.columns(4)
        for i, (col, label) in enumerate([
            ('GOALS', 'Goals per Game'), ('SHOTS ON TARGET', 'Shot Accuracy'),
            ('SUCCESSFUL PASSES', 'Pass Success Rate'), ('ATTK OUTCOME - SHOTS', 'Attacks to Shots')
        ]):
            plot_sparkline(row1[i], col, label)

        row2 = st.columns(4)
        for i, (col, label) in enumerate([
            ('CROSSES', 'Crosses'), ('FOULS', 'Fouls Committed'),
            ('PASSES PER ATTACK AVG.', 'Passes Per Attack'), ('SHOT BLOCKS', 'Shot Blocks')
        ]):
            plot_sparkline(row2[i], col, label)

    # GK Summary
    st.markdown("---")
    st.markdown("## GOALKEEPER REPORT")
    for _, gk_row in gk.iterrows():
        if gk_row['SA'] > 0 or gk_row['Sv'] > 0 or gk_row['GA'] > 0:
            gk_cols = st.columns(6)
            gk_cols[0].metric("GOALKEEPER", gk_row['Name'])
            gk_cols[1].metric("SHOTS AGAINST", int(gk_row['SA']))
            gk_cols[2].metric("GOALS AGAINST", int(gk_row['GA']))
            gk_cols[3].metric("SAVES", int(gk_row['Sv']))
            gk_cols[4].metric("SAVE %", f"{gk_row['Sv%']}%")
            gk_cols[5].metric("PIM", int(gk_row['PIM']))
            st.markdown("")


# ══════════════════════════════════════════════
# PAGE: PLUS/MINUS
# ══════════════════════════════════════════════
elif page == "Plus/Minus":
    st.markdown("""
    <h1 style="letter-spacing:3px;">PLUS / MINUS</h1>
    <p style="color:#777;">Who makes the team better when they're on the field? Game-by-game impact tracking.</p>
    """, unsafe_allow_html=True)

    pm_data = data['plusminus']
    game_dates = data['pm_game_dates']

    # Leaderboard
    st.markdown("## SEASON LEADERBOARD")
    pm_df = pd.DataFrame([{
        'Rank': i + 1, 'Name': p['Name'],
        '+': int(p['Plus']), '-': int(p['Minus']),
        'NET': int(p['Total']), 'Trend': p['Trend'],
    } for i, p in enumerate(sorted(pm_data, key=lambda x: x['Total'], reverse=True))])
    st.dataframe(pm_df, use_container_width=True, hide_index=True, height=500)

    st.markdown("---")

    # Player Deep Dive
    st.markdown("## PLAYER DEEP DIVE")
    player_names = [p['Name'] for p in pm_data]
    selected_player = st.selectbox("Select Player", player_names, key="pm_player")
    player = next(p for p in pm_data if p['Name'] == selected_player)
    game_vals = player['game_values']
    while len(game_vals) < len(game_dates):
        game_vals.append(None)

    info_c1, info_c2, info_c3, info_c4 = st.columns(4)
    info_c1.metric("TOTAL PLUS", f"+{int(player['Plus'])}")
    info_c2.metric("TOTAL MINUS", int(player['Minus']))
    net = int(player['Total'])
    info_c3.metric("NET +/-", f"+{net}" if net > 0 else str(net),
                    delta=player['Trend'] if player['Trend'] not in ['-', '0', '0.0'] else None)
    games_played = sum(1 for v in game_vals if v is not None)
    info_c4.metric("GAMES TRACKED", games_played)

    # Compute valid values for streaks
    valid_dates = []
    valid_vals = []
    for d, v in zip(game_dates, game_vals):
        if v is not None:
            valid_dates.append(d)
            valid_vals.append(v)

    # Hot/Cold Streaks
    st.markdown("### HOT & COLD STREAKS")
    if valid_vals:
        streak_fig = go.Figure()
        streak_colors = []
        for v in game_vals:
            if v is None: streak_colors.append('#1a1a1a')
            elif v > 2: streak_colors.append('#0fb')
            elif v > 0: streak_colors.append(TEAL)
            elif v == 0: streak_colors.append('#555')
            elif v > -2: streak_colors.append('#cc6633')
            else: streak_colors.append('#cc3333')

        for i, (d, v, c) in enumerate(zip(game_dates, game_vals, streak_colors)):
            display_val = f"+{int(v)}" if v is not None and v > 0 else (str(int(v)) if v is not None else "DNP")
            streak_fig.add_trace(go.Bar(
                x=[d], y=[1], marker_color=c,
                text=display_val, textposition='inside',
                textfont=dict(color='white', size=9, family='Arial'),
                showlegend=False, hoverinfo='text',
                hovertext=f"{d}: {display_val}"
            ))
        streak_fig = styled_fig(streak_fig, 120)
        streak_fig.update_layout(
            barmode='group',
            yaxis=dict(showticklabels=False, showgrid=False),
            xaxis=dict(showgrid=False),
            margin=dict(l=10, r=10, t=10, b=30), bargap=0.05
        )
        st.plotly_chart(streak_fig, use_container_width=True)

        st.markdown(f"""
        <div class="streak-legend">
            <span><span style="color:#0fb;">&#9679;</span> Hot (+3+)</span>
            <span><span style="color:{TEAL};">&#9679;</span> Positive</span>
            <span><span style="color:#555;">&#9679;</span> Even</span>
            <span><span style="color:#cc6633;">&#9679;</span> Slight Neg</span>
            <span><span style="color:#cc3333;">&#9679;</span> Cold (-2-)</span>
            <span><span style="color:#333;">&#9679;</span> DNP</span>
        </div>
        """, unsafe_allow_html=True)

    # Team-Wide Comparison
    st.markdown("---")
    st.markdown("## TEAM-WIDE COMPARISON")
    all_pm = sorted(pm_data, key=lambda x: x['Total'], reverse=True)
    fig = go.Figure()
    fig.add_trace(go.Bar(
        x=[p['Name'] for p in all_pm],
        y=[p['Plus'] for p in all_pm],
        name='Goals For (On Field)', marker_color=TEAL, opacity=0.8
    ))
    fig.add_trace(go.Bar(
        x=[p['Name'] for p in all_pm],
        y=[-p['Minus'] for p in all_pm],
        name='Goals Against (On Field)', marker_color='#cc3333', opacity=0.8
    ))
    fig.add_hline(y=0, line_color="#fff", line_width=1)
    fig = styled_fig(fig, 420)
    fig.update_layout(
        barmode='relative', xaxis_tickangle=-40, yaxis_title='Goals Impact',
        legend=dict(orientation='h', y=1.1, x=0.5, xanchor='center')
    )
    st.plotly_chart(fig, use_container_width=True)


# ══════════════════════════════════════════════
# PAGE: HUDL TREND ANALYSIS
# ══════════════════════════════════════════════
elif page == "Hudl Trend Analysis":
    st.markdown("""
    <h1 style="letter-spacing:3px;">HUDL TREND ANALYSIS</h1>
    <p style="color:#777;">Track team progression across the season. Understand how the Ambush are evolving game by game.</p>
    """, unsafe_allow_html=True)

    hudl = data['hudl']
    ambush = hudl[hudl['Team'].str.contains('Ambush', case=False, na=False)].sort_values('Date').copy()
    opponents = hudl[~hudl['Team'].str.contains('Ambush', case=False, na=False)].sort_values('Date').copy()

    metric_categories = {
        'SHOOTING': ['GOALS', 'SHOTS', 'SHOTS ON TARGET', 'SHOTS OFF TARGET', 'SHOTS SAVED', 'SHOTS BLOCKED BY OPPONENT'],
        'PASSING': ['PASSES', 'SUCCESSFUL PASSES', 'UNSUCCESSFUL PASSES', 'PASSES PER ATTACK AVG.', 'PASS STRINGS OF 0', 'CROSSES'],
        'ATTACK ORIGINS': ['ATTACKS STARTING FROM DEF THIRD', 'ATTACKS STARTING FROM MID THIRD', 'ATTACKS STARTING FROM FINAL THIRD'],
        'ATTACK OUTCOMES': ['ATTACKS ENDING IN DEF THIRD', 'ATTACKS ENDING IN MID THIRD', 'ATTACKS ENDING IN FINAL THIRD'],
        'ATTACK RESULTS': ['ATTK OUTCOME - ENDS IN TRANSITION TO DEF', 'ATTK OUTCOME - SHOTS', 'ATTK OUTCOME - FOULS', 'ATTK OUTCOME - BALL OUT', 'ATTK OUTCOME - GK PICKUP'],
        'DEFENSIVE': ['SHOT BLOCKS', 'SAVES', 'GK THROWS', 'FOULS', 'ASSISTS']
    }

    available_metrics = [c for c in ambush.columns if c not in ['Date', 'Team', 'Date_str']]

    tab1, tab2, tab3 = st.tabs(["PROGRESSION TRENDS", "HEAD-TO-HEAD", "HEAT MAP"])

    with tab1:
        st.markdown("### AMBUSH PROGRESSION — SELECT METRICS")
        st.markdown(f'<p style="color:#666; font-size:0.8rem;">The <span style="color:{TEAL};">teal line</span> shows per-game performance. The <span style="color:#888;">dashed line</span> is the season average baseline.</p>', unsafe_allow_html=True)

        cat = st.selectbox("Metric Category", list(metric_categories.keys()), key="hudl_cat")
        cat_metrics = [m for m in metric_categories[cat] if any(c.strip() == m.strip() for c in ambush.columns)]
        selected_metrics = st.multiselect("Select Metrics to Display", cat_metrics, default=cat_metrics[:3], key="hudl_metrics")

        for metric in selected_metrics:
            real_col = None
            for c in ambush.columns:
                if c.strip() == metric.strip():
                    real_col = c
                    break
            if real_col:
                vals = ambush[real_col].values
                avg_val = np.mean(vals)
                fig = go.Figure()
                fig.add_trace(go.Scatter(
                    x=ambush['Date_str'], y=vals,
                    mode='lines+markers', name='Per Game',
                    line=dict(color=TEAL, width=2.5),
                    marker=dict(color=TEAL, size=8, line=dict(color='white', width=1)),
                    fill='tozeroy', fillcolor='rgba(7,121,136,0.1)'
                ))
                fig.add_hline(y=avg_val, line_dash="dash", line_color="#888",
                              annotation_text=f"Avg: {avg_val:.2f}",
                              annotation_position="top right",
                              annotation_font_color="#999")
                fig = styled_fig(fig, 320)
                fig.update_layout(
                    title=dict(text=metric.strip(), font=dict(family='Anton, Impact', size=13, color='#ccc')),
                    xaxis_title='Game Date', showlegend=False
                )
                st.plotly_chart(fig, use_container_width=True)

    with tab2:
        st.markdown("### AMBUSH VS OPPONENT — PER GAME COMPARISON")
        cat2 = st.selectbox("Metric Category", list(metric_categories.keys()), key="h2h_cat")
        cat2_metrics = [m for m in metric_categories[cat2] if any(c.strip() == m.strip() for c in ambush.columns)]
        h2h_metric = st.selectbox("Select Metric", cat2_metrics, key="h2h_metric")

        real_col_a, real_col_o = None, None
        for c in ambush.columns:
            if c.strip() == h2h_metric.strip(): real_col_a = c; break
        for c in opponents.columns:
            if c.strip() == h2h_metric.strip(): real_col_o = c; break

        if real_col_a and real_col_o:
            game_dates_h = sorted(hudl['Date'].unique())
            amb_vals, opp_vals, opp_names, date_labels = [], [], [], []
            for d in game_dates_h:
                a = ambush[ambush['Date'] == d]
                o = opponents[opponents['Date'] == d]
                if len(a) > 0 and len(o) > 0:
                    amb_vals.append(a[real_col_a].values[0])
                    opp_vals.append(o[real_col_o].values[0])
                    opp_names.append(o['Team'].values[0])
                    date_labels.append(pd.Timestamp(d).strftime('%m/%d'))

            fig = go.Figure()
            fig.add_trace(go.Bar(x=date_labels, y=amb_vals, name='Ambush', marker_color=TEAL, opacity=0.9))
            fig.add_trace(go.Bar(x=date_labels, y=opp_vals, name='Opponent', marker_color='#cc3333', opacity=0.7))
            for i, name in enumerate(opp_names):
                short = name.replace('Saint Louis ', '').replace('Kansas City ', 'KC ').replace('Milwaukee ', 'MIL ').replace('Empire ', 'EMP ').replace('Tacoma ', 'TAC ')
                fig.add_annotation(x=date_labels[i], y=max(amb_vals[i], opp_vals[i]) * 1.1,
                                   text=short, showarrow=False, font=dict(size=7, color='#777'))
            fig = styled_fig(fig, 420)
            fig.update_layout(barmode='group',
                              title=dict(text=h2h_metric.strip(), font=dict(family='Anton, Impact', size=13, color='#ccc')),
                              legend=dict(orientation='h', y=1.12, x=0.5, xanchor='center'))
            st.plotly_chart(fig, use_container_width=True)

            diffs = [a - o for a, o in zip(amb_vals, opp_vals)]
            diff_colors = [TEAL if d >= 0 else '#cc3333' for d in diffs]
            fig2 = go.Figure()
            fig2.add_trace(go.Bar(x=date_labels, y=diffs, marker_color=diff_colors,
                                  text=[f"+{d:.2f}" if d > 0 else f"{d:.2f}" for d in diffs],
                                  textposition='outside', textfont=dict(size=8, color='#ccc')))
            fig2.add_hline(y=0, line_color="#555", line_dash="dash")
            fig2 = styled_fig(fig2, 280)
            fig2.update_layout(title=dict(text="DIFFERENTIAL (Ambush - Opponent)", font=dict(family='Anton, Impact', size=11, color='#999')))
            st.plotly_chart(fig2, use_container_width=True)

    with tab3:
        st.markdown("### PERFORMANCE HEAT MAP — IMPROVEMENT / DECLINE")
        st.markdown('<p style="color:#666; font-size:0.8rem;">Green = above season average, Red = below. Intensity shows magnitude of deviation.</p>', unsafe_allow_html=True)

        cat3 = st.selectbox("Metric Category", list(metric_categories.keys()), key="hm_cat")
        cat3_metrics = [m for m in metric_categories[cat3] if any(c.strip() == m.strip() for c in ambush.columns)]

        if cat3_metrics:
            z_data, y_labels = [], []
            x_labels = ambush['Date_str'].tolist()
            for metric in cat3_metrics:
                real_col = None
                for c in ambush.columns:
                    if c.strip() == metric.strip(): real_col = c; break
                if real_col:
                    vals = ambush[real_col].values.astype(float)
                    avg = np.mean(vals)
                    std = np.std(vals) if np.std(vals) > 0 else 1
                    z_data.append(((vals - avg) / std).tolist())
                    y_labels.append(metric.strip()[:30])

            if z_data:
                fig = go.Figure(data=go.Heatmap(
                    z=z_data, x=x_labels, y=y_labels,
                    colorscale=[[0, '#cc3333'], [0.35, '#661a1a'], [0.5, '#1a1a1a'], [0.65, TEAL_DARK], [1.0, '#0fb']],
                    zmid=0, text=[[f"{v:.2f}" for v in row] for row in z_data],
                    texttemplate="%{text}", textfont=dict(size=8, color='white'),
                    hovertemplate='%{y}<br>%{x}<br>Z-Score: %{z:.2f}<extra></extra>',
                    colorbar=dict(title='Std Dev', tickvals=[-2, -1, 0, 1, 2],
                                  ticktext=['Below -2s', '-1s', 'Average', '+1s', 'Above +2s'],
                                  tickfont=dict(color='#999'))
                ))
                fig = styled_fig(fig, max(300, len(cat3_metrics) * 55))
                fig.update_layout(
                    title=dict(text="Z-Score Heat Map (vs Season Average)", font=dict(family='Anton, Impact', size=11, color='#ccc')),
                    xaxis_title='Game Date', yaxis=dict(autorange='reversed')
                )
                st.plotly_chart(fig, use_container_width=True)

        st.markdown("---")
        st.markdown("### 3-GAME ROLLING AVERAGE")
        roll_metric = st.selectbox("Select Metric", cat3_metrics, key="roll_metric")
        real_col = None
        for c in ambush.columns:
            if c.strip() == roll_metric.strip(): real_col = c; break
        if real_col:
            vals = ambush[real_col].values.astype(float)
            rolling = pd.Series(vals).rolling(3, min_periods=1).mean().values
            fig = go.Figure()
            fig.add_trace(go.Scatter(x=ambush['Date_str'].values, y=vals, mode='markers', name='Per Game',
                                     marker=dict(color=TEAL, size=8, line=dict(color='white', width=1))))
            fig.add_trace(go.Scatter(x=ambush['Date_str'].values, y=rolling, mode='lines', name='3-Game Rolling Avg',
                                     line=dict(color='#fff', width=2.5)))
            avg = np.mean(vals)
            fig.add_hline(y=avg, line_dash="dot", line_color="#555",
                          annotation_text=f"Season Avg: {avg:.2f}", annotation_font_color="#777")
            fig = styled_fig(fig, 350)
            fig.update_layout(legend=dict(orientation='h', y=1.1, x=0.5, xanchor='center'),
                              title=dict(text=roll_metric.strip(), font=dict(family='Anton, Impact', size=11, color='#ccc')))
            st.plotly_chart(fig, use_container_width=True)


# ══════════════════════════════════════════════
# PAGE: PLAYER STATS
# ══════════════════════════════════════════════
elif page == "Player Stats":
    st.markdown("""
    <h1 style="letter-spacing:3px;">PLAYER STATS CENTER</h1>
    <p style="color:#777;">Full roster breakdown with advanced per-game analytics.</p>
    """, unsafe_allow_html=True)

    players = data['players']

    tab1, tab2, tab3 = st.tabs(["FULL ROSTER", "PLAYER PROFILE", "GOALKEEPERS"])

    with tab1:
        st.markdown("## SEASON STATS — FULL ROSTER")
        pos_filter = st.multiselect("Filter by Position", ['F', 'M', 'D'], default=['F', 'M', 'D'], key="pos_filter")
        sort_by = st.selectbox("Sort By", ['Pts', 'G', 'A', 'SOG', 'GP', 'Pts/GP', 'G/GP', 'Sh/GP'], key="sort_stat")
        filtered = players[players['Pos'].isin(pos_filter)].sort_values(sort_by, ascending=False).reset_index(drop=True)
        filtered.index = filtered.index + 1
        display_cols = ['#', 'Name', 'Pos', 'GP', 'G', 'A', 'Pts', 'Pts/GP', 'G/GP', 'SOG', 'Sh', 'SOG%', 'Sh/GP', 'BS', 'F', 'PIM', 'YC', 'RC']
        st.dataframe(filtered[display_cols], use_container_width=True, hide_index=False, height=600)

        st.markdown("---")
        st.markdown("## PRODUCTION BY POSITION")
        pos_stats = players.groupby('Pos').agg({'G': 'sum', 'A': 'sum', 'Pts': 'sum', 'F': 'sum', 'BS': 'sum'}).reset_index()
        pos_map = {'F': 'Forwards', 'M': 'Midfielders', 'D': 'Defenders'}
        pos_stats['Position'] = pos_stats['Pos'].map(pos_map)
        fig = go.Figure()
        fig.add_trace(go.Bar(x=pos_stats['Position'], y=pos_stats['G'], name='Goals', marker_color=TEAL))
        fig.add_trace(go.Bar(x=pos_stats['Position'], y=pos_stats['A'], name='Assists', marker_color='#444'))
        fig = styled_fig(fig, 350)
        fig.update_layout(barmode='group', legend=dict(orientation='h', y=1.1, x=0.5, xanchor='center'))
        st.plotly_chart(fig, use_container_width=True)

    with tab2:
        st.markdown("## PLAYER PROFILE")
        player_name = st.selectbox("Select Player", players.sort_values('Pts', ascending=False)['Name'].tolist(), key="player_profile")
        p = players[players['Name'] == player_name].iloc[0]

        c1, c2, c3, c4, c5 = st.columns(5)
        c1.metric("NUMBER", f"#{int(p['#'])}")
        c2.metric("POSITION", p['Pos'])
        c3.metric("GAMES PLAYED", int(p['GP']))
        c4.metric("POINTS", int(p['Pts']))
        c5.metric("PTS/GAME", p['Pts/GP'])

        st.markdown("### OFFENSIVE OUTPUT")
        o1, o2, o3, o4 = st.columns(4)
        o1.metric("GOALS", int(p['G']), delta=f"{p['G/GP']} per game")
        o2.metric("ASSISTS", int(p['A']))
        o3.metric("SHOTS", int(p['Sh']), delta=f"{p['Sh/GP']} per game")
        o4.metric("SOG %", f"{p['SOG%']}%")

        st.markdown("### DEFENSIVE & DISCIPLINE")
        d1, d2, d3, d4, d5 = st.columns(5)
        d1.metric("BLOCKS", int(p['BS']))
        d2.metric("FOULS", int(p['F']))
        d3.metric("BLUE CARDS", int(p['BC']))
        d4.metric("YELLOW CARDS", int(p['YC']))
        d5.metric("PIM", int(p['PIM']))

        # Radar chart — per game stats
        st.markdown("### PLAYER RADAR — PER GAME")
        gp = p['GP'] if p['GP'] > 0 else 1
        categories = ['G/GP', 'A/GP', 'SOG/GP', 'BS/GP', 'F/GP']
        players_pg = players.copy()
        players_pg['_gp'] = players_pg['GP'].replace(0, 1)
        max_vals = [
            (players_pg['G'] / players_pg['_gp']).max(),
            (players_pg['A'] / players_pg['_gp']).max(),
            (players_pg['SOG'] / players_pg['_gp']).max(),
            (players_pg['BS'] / players_pg['_gp']).max(),
            (players_pg['F'] / players_pg['_gp']).max(),
        ]
        player_vals = [p['G']/gp, p['A']/gp, p['SOG']/gp, p['BS']/gp, p['F']/gp]
        normalized = [pv / mv if mv > 0 else 0 for pv, mv in zip(player_vals, max_vals)]

        fig = go.Figure()
        fig.add_trace(go.Scatterpolar(
            r=normalized + [normalized[0]],
            theta=categories + [categories[0]],
            fill='toself', fillcolor='rgba(7,121,136,0.3)',
            line=dict(color=TEAL, width=2),
            marker=dict(color=TEAL, size=6)
        ))
        fig = styled_fig(fig, 380)
        fig.update_layout(
            polar=dict(
                bgcolor='rgba(17,17,17,0.8)',
                radialaxis=dict(visible=True, range=[0, 1], gridcolor='#333', tickfont=dict(color='#555')),
                angularaxis=dict(gridcolor='#333', tickfont=dict(color='#aaa', family='Anton, Impact', size=10))
            )
        )
        st.plotly_chart(fig, use_container_width=True)

    with tab3:
        st.markdown("## GOALKEEPER STATS")
        gk = data['goalkeepers']
        for _, row in gk.iterrows():
            st.markdown(f"### #{int(row['#'])} {row['Name']}")
            g1, g2, g3, g4, g5 = st.columns(5)
            g1.metric("SHOTS AGAINST", int(row['SA']))
            g2.metric("GOALS AGAINST", int(row['GA']))
            g3.metric("SAVES", int(row['Sv']))
            g4.metric("SAVE %", f"{row['Sv%']}%")
            g5.metric("PIM", int(row['PIM']))
            if row['SA'] > 0:
                fig = go.Figure(data=[go.Pie(
                    labels=['Saves', 'Goals Against'],
                    values=[row['Sv'], row['GA']],
                    hole=0.65, marker=dict(colors=[TEAL, '#cc3333']),
                    textinfo='label+value', textfont=dict(color='white', size=12)
                )])
                fig = styled_fig(fig, 280)
                fig.update_layout(
                    annotations=[dict(text=f"{row['Sv%']}%", x=0.5, y=0.5,
                                      font_size=20, font_color='white',
                                      font_family='Anton, Impact', showarrow=False)]
                )
                st.plotly_chart(fig, use_container_width=True)
            st.markdown("---")


# ══════════════════════════════════════════════
# PAGE: CONTACT
# ══════════════════════════════════════════════
elif page == "Contact":
    st.markdown(f"""
    <div style="max-width:700px; margin:2rem auto; text-align:center;">
        <h1 style="letter-spacing:3px;">CONTACT</h1>
        <div style="background:{CARD_BG}; border:1px solid {CARD_BORDER}; border-top:3px solid {TEAL}; border-radius:8px; padding:32px; margin-top:1.5rem;">
            <h2 style="margin-bottom:4px; color:#fff !important;">DYLLAN HAGGARD</h2>
            <p style="color:{TEAL}; font-family:'Anton',Impact,sans-serif; font-size:0.85rem; letter-spacing:2px; margin-top:0;">DATA ANALYST  |  VIDEO ANALYST  |  SOCCER ANALYTICS</p>
            <div style="border-top:1px solid #222; margin:20px 0; padding-top:20px;">
                <p style="color:#bbb; font-size:0.85rem; line-height:1.7;">
                    Data Analyst II at Stifel Bank & Trust and Video/Data Analyst for the St. Louis Ambush in the Major Arena Soccer League.
                    Passionate about bridging the gap between data science, video analysis, and coaching strategy to elevate the game at every level.
                </p>
            </div>
            <div style="margin-top:24px; display:flex; justify-content:center; gap:16px; flex-wrap:wrap;">
                <a href="mailto:dyllanleehaggard@gmail.com" style="
                    display:inline-block; background:{TEAL}; color:#fff; padding:10px 24px;
                    border-radius:6px; text-decoration:none; font-family:'Anton',Impact,sans-serif;
                    font-size:0.8rem; letter-spacing:1px; text-transform:uppercase;
                ">EMAIL ME</a>
                <a href="https://www.linkedin.com/in/dyllan-haggard/" target="_blank" style="
                    display:inline-block; background:transparent; color:{TEAL}; padding:10px 24px;
                    border-radius:6px; text-decoration:none; font-family:'Anton',Impact,sans-serif;
                    font-size:0.8rem; letter-spacing:1px; text-transform:uppercase;
                    border:1px solid {TEAL};
                ">LINKEDIN</a>
            </div>
        </div>
        <p style="color:#444; font-size:0.65rem; margin-top:2rem; letter-spacing:1px;">MASL 2025-26 SEASON  |  ST. LOUIS AMBUSH  |  COACHING INTELLIGENCE PLATFORM</p>
    </div>
    """, unsafe_allow_html=True)
